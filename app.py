"""
Financial Modeling Automation Studio
Showcasing PE/VC Excel automation with Python
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import io
from pathlib import Path

# Page config
st.set_page_config(
    page_title="Financial Modeling Automation",
    page_icon="📊",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        font-weight: 700;
        text-align: center;
        margin-bottom: 1rem;
        background: linear-gradient(120deg, #1e3a8a, #3b82f6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .sub-header {
        font-size: 1.2rem;
        text-align: center;
        color: #64748b;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown('<h1 class="main-header">📊 Financial Modeling Automation Studio</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">PE/VC Excel Models + Python Automation | MBA + Tech Stack</p>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://img.icons8.com/clouds/200/000000/money-bag.png", width=150)
    st.markdown("### 🎯 About This Project")
    st.markdown("""
    This app demonstrates:
    - **VC Cap Table Modeling**
    - **LP/GP Fund Economics**
    - **Python + openpyxl Automation**
    - **Interactive Analysis**
    
    **Tech Stack:**
    - Python (openpyxl, pandas)
    - Streamlit
    - Plotly
    - Excel automation
    """)
    
    st.markdown("---")
    st.markdown("### 👨‍💼 Built by")
    st.markdown("**MBA | PE/VC Professional**")
    st.markdown("Combining finance expertise with technical automation")

# Tabs
tab1, tab2, tab3 = st.tabs(["🏢 VC Cap Table", "💰 LP/GP Fund Model", "🔧 Excel Automation"])

with tab1:
    st.header("VC Capitalization Table Model")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("""
        ### Overview
        This model demonstrates a complete startup cap table across multiple funding rounds:
        - **Pre-money** structure with founders and option pool
        - **Seed Round** ($2M on $8M pre)
        - **Series A** ($10M on $40M pre)
        - **Series B** ($30M on $120M pre)
        """)
        
        # Sample assumptions
        st.markdown("### Key Assumptions")
        assumptions_df = pd.DataFrame({
            'Round': ['Initial', 'Seed', 'Series A', 'Series B'],
            'Pre-Money Val': ['$8M', '$8M', '$40M', '$120M'],
            'Investment': ['-', '$2M', '$10M', '$30M'],
            'Option Pool': ['12.5%', '10%', '10%', '10%']
        })
        st.dataframe(assumptions_df, use_container_width=True)
    
    with col2:
        st.markdown("### Key Metrics")
        st.metric("Total Capital Raised", "$42M")
        st.metric("Post-Series B Valuation", "$150M")
        st.metric("Founder Dilution", "~35%")
    
    # Ownership evolution chart
    st.markdown("### Ownership Evolution")
    
    # Sample data for visualization
    rounds = ['Pre-Money', 'Post-Seed', 'Post-Series A', 'Post-Series B']
    founders_pct = [87.5, 70.0, 56.0, 44.8]
    employees_pct = [12.5, 10.0, 10.0, 10.0]
    seed_pct = [0, 20.0, 16.0, 12.8]
    seriesa_pct = [0, 0, 18.0, 14.4]
    seriesb_pct = [0, 0, 0, 18.0]
    
    fig = go.Figure()
    fig.add_trace(go.Bar(name='Founders', x=rounds, y=founders_pct, marker_color='#3b82f6'))
    fig.add_trace(go.Bar(name='Employees', x=rounds, y=employees_pct, marker_color='#fbbf24'))
    fig.add_trace(go.Bar(name='Seed', x=rounds, y=seed_pct, marker_color='#10b981'))
    fig.add_trace(go.Bar(name='Series A', x=rounds, y=seriesa_pct, marker_color='#8b5cf6'))
    fig.add_trace(go.Bar(name='Series B', x=rounds, y=seriesb_pct, marker_color='#ef4444'))
    
    fig.update_layout(
        barmode='stack',
        title='Ownership % by Round',
        xaxis_title='Funding Round',
        yaxis_title='Ownership %',
        height=500
    )
    st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("### 📄 Features")
    st.markdown("""
    - **Dilution calculations** across multiple rounds
    - **Option pool mechanics** (pre/post-money)
    - **Price per share** tracking
    - **Automatic formula updates** via openpyxl
    - **Color-coded** stakeholder visualization
    """)

with tab2:
    st.header("LP/GP Venture Fund Economics Model")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("### Fund Structure")
        st.markdown("""
        **Fund Size:** $150M  
        **Management Fee:** 2% (years 1-5), 1.5% (years 6-10)  
        **Carry:** 20% of profits  
        **Investment Period:** 5 years  
        **Fund Life:** 10 years  
        """)
        
        # KPIs
        st.markdown("### Target Performance")
        kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
        with kpi_col1:
            st.metric("Target MOIC", "2.8x")
        with kpi_col2:
            st.metric("Target IRR", "20%")
        with kpi_col3:
            st.metric("GP Carry", "$84M")
    
    with col2:
        st.markdown("### Capital Deployment")
        # Sample data
        years = list(range(1, 11))
        capital_called = [30, 30, 30, 30, 30, 0, 0, 0, 0, 0]
        mgmt_fees = [3, 3, 3, 3, 3, 2.25, 2.25, 2.25, 2.25, 2.25]
        
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(name='Capital Called', x=years, y=capital_called, marker_color='#3b82f6'))
        fig2.add_trace(go.Bar(name='Management Fees', x=years, y=mgmt_fees, marker_color='#ef4444'))
        
        fig2.update_layout(
            title='Annual Capital Deployment ($M)',
            xaxis_title='Year',
            yaxis_title='Amount ($M)',
            height=300
        )
        st.plotly_chart(fig2, use_container_width=True)
    
    # Distribution waterfall
    st.markdown("### Distribution Waterfall")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Exit proceeds by year
        years_exit = list(range(6, 11))
        exit_proceeds = [42, 84, 126, 126, 42]  # Sample data
        
        fig3 = px.bar(x=years_exit, y=exit_proceeds, 
                      title='Exit Proceeds by Year ($M)',
                      labels={'x': 'Year', 'y': 'Proceeds ($M)'},
                      color_discrete_sequence=['#10b981'])
        fig3.update_layout(height=350)
        st.plotly_chart(fig3, use_container_width=True)
    
    with col2:
        # LP vs GP distributions
        cumulative_years = list(range(1, 11))
        lp_cumulative = [0, 0, 0, 0, 0, 33.6, 100.8, 201.6, 302.4, 336]
        gp_cumulative = [0, 0, 0, 0, 0, 8.4, 25.2, 50.4, 75.6, 84]
        
        fig4 = go.Figure()
        fig4.add_trace(go.Scatter(name='LP', x=cumulative_years, y=lp_cumulative, 
                                  mode='lines+markers', line=dict(color='#3b82f6', width=3)))
        fig4.add_trace(go.Scatter(name='GP', x=cumulative_years, y=gp_cumulative, 
                                  mode='lines+markers', line=dict(color='#8b5cf6', width=3)))
        
        fig4.update_layout(
            title='Cumulative Distributions ($M)',
            xaxis_title='Year',
            yaxis_title='Amount ($M)',
            height=350
        )
        st.plotly_chart(fig4, use_container_width=True)
    
    st.markdown("### 📊 Model Features")
    st.markdown("""
    - **Capital call schedules** with management fee calculations
    - **Carry waterfall** (20% over capital returned)
    - **IRR calculations** for LP returns
    - **DPI and MOIC metrics**
    - **Exit scenario modeling**
    """)

with tab3:
    st.header("Excel Automation with Python")
    
    st.markdown("""
    ### 🔧 Technical Implementation
    
    Both models are **fully automated** using Python's `openpyxl` library, demonstrating:
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        #### Automation Features
        ✅ **Dynamic formula generation**  
        ✅ **Multi-sheet workbook creation**  
        ✅ **Cell styling & formatting**  
        ✅ **Conditional formatting**  
        ✅ **Chart generation**  
        ✅ **Data validation**  
        """)
    
    with col2:
        st.markdown("""
        #### Use Cases
        💼 **Rapid scenario modeling**  
        💼 **Template standardization**  
        💼 **Error reduction**  
        💼 **Audit trail automation**  
        💼 **Report generation**  
        💼 **Integration with data pipelines**  
        """)
    
    st.markdown("---")
    
    # Code examples
    st.markdown("### 📝 Sample Code: Cap Table Automation")
    
    with st.expander("View Python Code - Creating Cap Table"):
        st.code("""
# Example: Automated cap table generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_cap_table(pre_money_val, investment, option_pool_pct):
    wb = Workbook()
    ws = wb.active
    
    # Header styling
    header_fill = PatternFill(start_color="366092", 
                              end_color="366092", 
                              fill_type='solid')
    
    # Create assumptions sheet
    ws['A1'] = 'Pre-Money Valuation'
    ws['B1'] = pre_money_val
    ws['B1'].number_format = '$#,##0'
    
    # Calculate dilution
    post_money = pre_money_val + investment
    investor_pct = investment / post_money
    
    # Create cap table with formulas
    ws['A5'] = 'Stakeholder'
    ws['B5'] = 'Shares'
    ws['C5'] = 'Ownership %'
    
    # Dynamic formulas
    ws['C6'] = '=B6/SUM($B$6:$B$10)'
    
    return wb

# Run automation
cap_table = create_cap_table(8000000, 2000000, 0.10)
cap_table.save('automated_cap_table.xlsx')
        """, language='python')
    
    with st.expander("View Python Code - LP/GP Fund Model"):
        st.code("""
# Example: Automated LP/GP fund model
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

def create_fund_model(fund_size, carry_pct, mgmt_fee):
    wb = Workbook()
    ws_capital = wb.create_sheet("Capital Calls")
    
    # Generate capital call schedule
    for year in range(1, 11):
        row = year + 1
        ws_capital[f'A{row}'] = year
        
        # Management fee formula
        ws_capital[f'C{row}'] = (
            f'=IF(A{row}<=5,{mgmt_fee},{mgmt_fee*0.75})'
            f'*Assumptions!B1'
        )
        
        # Net invested
        ws_capital[f'D{row}'] = f'=B{row}-C{row}'
    
    # Create distributions sheet with carry calculation
    ws_dist = wb.create_sheet("Distributions")
    
    # Carry waterfall
    ws_dist['D2'] = f'=MAX(0,(B2-SUM(Capital!D:D))*{carry_pct})'
    
    # Add charts
    chart = LineChart()
    chart.title = "Cumulative Distributions"
    data = Reference(ws_dist, min_col=2, min_row=1, max_row=11)
    chart.add_data(data, titles_from_data=True)
    ws_dist.add_chart(chart, "F2")
    
    return wb

# Run automation
fund_model = create_fund_model(150000000, 0.20, 0.02)
fund_model.save('automated_fund_model.xlsx')
        """, language='python')
    
    st.markdown("---")
    
    # Upload section
    st.markdown("### 📤 Try It Yourself")
    st.markdown("Upload your own Excel financial model to extract and visualize key metrics")
    
    uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx', 'xls'])
    
    if uploaded_file:
        try:
            # Read Excel
            df = pd.read_excel(uploaded_file, sheet_name=0)
            st.success(f"✅ Loaded {uploaded_file.name} successfully!")
            
            st.markdown("#### Preview")
            st.dataframe(df.head(), use_container_width=True)
            
            st.markdown("#### Basic Statistics")
            st.write(df.describe())
            
        except Exception as e:
            st.error(f"Error reading file: {str(e)}")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #64748b;'>
    <p><strong>Financial Modeling Automation Studio</strong></p>
    <p>Production-grade financial models • Automated with Python</p>
    <p>Tech: Streamlit • openpyxl • Plotly</p>
</div>
""", unsafe_allow_html=True)

